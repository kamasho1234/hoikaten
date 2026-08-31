"""hoikaten.com の対応状況を、全国市区町村の一覧にして Excel に書き出す

実行: python scripts/coverage-report.py <出力先.xlsx>

## なぜ作ったか
「どの自治体まで手が回っているか」を人口つきで見たいとき、
リポジトリの中身（点数・空き状況・お金記事）を数えるだけでは
**まだ手を付けていない自治体が表に出てこない**。
全国の市区町村を土台にして、そこへ対応状況を重ねる。

## 人口の出どころ
総務省「住民基本台帳に基づく人口、人口動態及び世帯数」の市区町村別（総計）。
一覧ページから最新年の Excel を毎回たどるので、年が変わっても直さなくてよい。
https://www.soumu.go.jp/main_sosiki/jichi_gyousei/daityo/jinkou_jinkoudoutai-setaisuu.html

## 行の単位
市・町・村・特別区。政令指定都市は市単位（行政区の行は落とす）。
都道府県の合計行と郡の合計行も落とす。
"""

import glob
import io
import json
import os
import re
import sys
import tempfile
import urllib.request

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO = os.getcwd()
INDEX_URL = "https://www.soumu.go.jp/main_sosiki/jichi_gyousei/daityo/jinkou_jinkoudoutai-setaisuu.html"
UA = "Mozilla/5.0 (compatible; hoikaten/1.0; +https://hoikaten.com)"

# 総務省の表とサイト側で字が違うことがある（四條畷市／四条畷市、須惠町／須恵町 など）
VARIANTS = str.maketrans({
    "條": "条", "惠": "恵", "澤": "沢", "龍": "竜", "邊": "辺", "邉": "辺",
    "嶋": "島", "榮": "栄", "濱": "浜", "舘": "館", "ヶ": "ケ", "﨑": "崎", "德": "徳",
})


def fail(message):
    raise SystemExit(f"[中断] {message}")


def norm(text):
    return (text or "").translate(VARIANTS)


def fetch(url):
    req = urllib.request.Request(url, headers={"User-Agent": UA})
    return urllib.request.urlopen(req, timeout=120).read()


def download_population():
    """総務省の一覧ページから、市区町村別（総計）の Excel を落としてくる"""
    raw = fetch(INDEX_URL)
    html = None
    for enc in ("cp932", "utf-8", "euc_jp"):
        try:
            html = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if html is None:
        fail("総務省のページの文字コードを判別できませんでした")

    hits = []
    for m in re.finditer(r'<a[^>]+href="([^"]+\.xlsx)"[^>]*>([\s\S]{0,150}?)</a>', html):
        label = re.sub(r"<[^>]*>", "", m.group(2)).strip()
        if "市区町村別" in label and "総計" in label and "年齢階級" not in label:
            hits.append((m.group(1), label))
    if not hits:
        fail("市区町村別（総計）のExcelが総務省のページで見つかりません")

    url, label = hits[0]
    if not url.startswith("http"):
        url = "https://www.soumu.go.jp" + url
    print(f"人口: {label}")
    print(f"     {url}")
    path = os.path.join(tempfile.mkdtemp(), "jinko.xlsx")
    with open(path, "wb") as f:
        f.write(fetch(url))
    return path, label


def load_municipalities(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    out = []
    for r in ws.iter_rows(min_row=7, values_only=True):
        code, pref, name = r[0], r[1], r[2]
        if not code or not pref or name in (None, "", "-"):
            continue
        name = str(name).strip()
        if name.endswith("郡"):
            continue
        if re.match(r".+市.+区$", name):
            continue
        # 「西多摩郡瑞穂町」→「瑞穂町」。ただし「大和郡山市」「小郡市」は郡が名前の一部なので、
        # 郡を外すのは残りが町か村で終わるときだけにする
        m = re.match(r"^(.+郡)(.+[町村])$", name)
        out.append({
            "code": str(code),
            "pref": str(pref).strip(),
            "name": m.group(2) if m else name,
            "pop": r[5] or 0,
        })
    wb.close()
    if len(out) < 1700:
        fail(f"市区町村の行が {len(out)} 件しか取れませんでした（表の形が変わった可能性）")
    return out


def load_coverage():
    """リポジトリから、点数・空き状況・お金記事のそれぞれの対応自治体を集める"""
    sim = {}
    for p in glob.glob(os.path.join(REPO, "src/lib/data/*.ts")):
        slug = os.path.basename(p)[:-3]
        if slug == "index":
            continue
        t = io.open(p, encoding="utf-8").read(4000)
        n = re.search(r"name:\s*'([^']+)'", t) or re.search(r'name:\s*"([^"]+)"', t)
        pr = re.search(r"prefecture:\s*'([^']+)'", t) or re.search(r'prefecture:\s*"([^"]+)"', t)
        sim[slug] = (n.group(1) if n else "", pr.group(1) if pr else "")

    money = set()
    for p in glob.glob(os.path.join(REPO, "src/lib/insurance/city-*.ts")):
        t = io.open(p, encoding="utf-8").read()
        money.update(re.findall(r'citySlug:\s*"([a-z0-9\-]+)"', t))

    s = io.open(os.path.join(REPO, "src/lib/vacancy/index.ts"), encoding="utf-8").read()
    start = s.index("const registry")
    end = s.index("\n};\n", start)
    vac = set(re.findall(
        r'^\s+"?([a-z0-9\-]+)"?:\s+(?:\w+ as unknown as VacancyDataset|withWebsites)',
        s[start:end], re.M))

    # 点数ページが無い「空き状況だけ」の自治体は、JSONから名前と県を取る
    vac_only = {}
    for slug in sorted(vac - set(sim)):
        d = json.load(io.open(os.path.join(REPO, f"src/lib/vacancy/{slug}.json"), encoding="utf-8"))
        vac_only[slug] = (d.get("municipalityName", ""), d.get("prefecture", ""))
    return sim, money, vac, vac_only


def build_rows(munis, sim, money, vac, vac_only):
    by_key = {}
    by_name = {}   # 県を持たないデータ（空き状況だけの自治体）の逃げ道
    for slug, (name, pref) in list(sim.items()) + list(vac_only.items()):
        if pref:
            by_key.setdefault((norm(pref), norm(name)), slug)
        else:
            by_name.setdefault(norm(name), slug)

    used = set()
    rows = []
    for m in munis:
        slug = by_key.get((norm(m["pref"]), norm(m["name"])))
        if not slug:
            # 同名の市が別の県にもある（府中市は東京都と広島県）ので、
            # 名前だけの突き合わせは県を持たないデータに限り、しかも1回だけ使う
            cand = by_name.get(norm(m["name"]))
            if cand and cand not in used:
                slug = cand
        if slug:
            used.add(slug)
        row = {
            "code": m["code"], "pref": m["pref"], "name": m["name"], "pop": m["pop"],
            "slug": slug or "",
            "sim": "○" if slug in sim else "",
            "vac": "○" if slug and slug in vac else "",
            "money": "○" if slug and slug in money else "",
        }
        row["done"] = sum(1 for k in ("sim", "vac", "money") if row[k] == "○")
        rows.append(row)

    unmatched = [(slug, name, pref) for slug, (name, pref) in
                 list(sim.items()) + list(vac_only.items()) if slug not in used]
    return rows, unmatched


def write_book(out_path, rows, unmatched, source_label):
    wb = openpyxl.Workbook()
    head_fill = PatternFill("solid", fgColor="1F3864")
    head_font = Font(color="FFFFFF", bold=True)
    ok_fill = PatternFill("solid", fgColor="E2F0D9")
    part_fill = PatternFill("solid", fgColor="FFF2CC")
    none_fill = PatternFill("solid", fgColor="FBE5E5")

    def head(ws, cols, widths):
        ws.append(cols)
        for c in ws[1]:
            c.fill = head_fill
            c.font = head_font
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws.freeze_panes = "A2"
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    ws = wb.active
    ws.title = "全自治体"
    head(ws, ["団体コード", "都道府県", "市区町村", "人口", "シミュレーター", "空き状況", "お金記事", "対応数", "slug"],
         [12, 11, 20, 11, 14, 11, 11, 9, 18])
    # 人口の多い順。上から順に穴を埋めていく使い方をしているので、この並びを既定にする
    for r in sorted(rows, key=lambda x: (-x["pop"], x["code"])):
        ws.append([r["code"], r["pref"], r["name"], r["pop"], r["sim"], r["vac"], r["money"], r["done"], r["slug"]])
        fill = ok_fill if r["done"] == 3 else (none_fill if r["done"] == 0 else part_fill)
        for c in ws[ws.max_row]:
            c.fill = fill
    ws.auto_filter.ref = f"A1:I{ws.max_row}"
    for i in range(2, ws.max_row + 1):
        ws.cell(row=i, column=4).number_format = "#,##0"

    ws2 = wb.create_sheet("都道府県別")
    head(ws2, ["都道府県", "市区町村数", "シミュレーター", "空き状況", "お金記事", "3つとも", "未着手",
               "人口計", "カバー人口（点数）", "人口カバー率"],
         [12, 12, 14, 11, 11, 10, 10, 14, 18, 13])
    order = sorted({r["pref"] for r in rows}, key=lambda x: min(r["code"] for r in rows if r["pref"] == x))
    for p in order:
        g = [r for r in rows if r["pref"] == p]
        pop = sum(r["pop"] for r in g)
        covered = sum(r["pop"] for r in g if r["sim"] == "○")
        ws2.append([p, len(g),
                    sum(1 for r in g if r["sim"] == "○"),
                    sum(1 for r in g if r["vac"] == "○"),
                    sum(1 for r in g if r["money"] == "○"),
                    sum(1 for r in g if r["done"] == 3),
                    sum(1 for r in g if r["done"] == 0),
                    pop, covered, round(covered / pop, 4) if pop else 0])
    pop = sum(r["pop"] for r in rows)
    covered = sum(r["pop"] for r in rows if r["sim"] == "○")
    ws2.append(["合計", len(rows),
                sum(1 for r in rows if r["sim"] == "○"),
                sum(1 for r in rows if r["vac"] == "○"),
                sum(1 for r in rows if r["money"] == "○"),
                sum(1 for r in rows if r["done"] == 3),
                sum(1 for r in rows if r["done"] == 0),
                pop, covered, round(covered / pop, 4) if pop else 0])
    for c in ws2[ws2.max_row]:
        c.font = Font(bold=True)
    for i in range(2, ws2.max_row + 1):
        ws2.cell(row=i, column=10).number_format = "0.0%"
        for col in (8, 9):
            ws2.cell(row=i, column=col).number_format = "#,##0"

    ws3 = wb.create_sheet("未着手（人口順）")
    head(ws3, ["順位", "都道府県", "市区町村", "人口", "団体コード"], [7, 11, 20, 12, 12])
    for i, r in enumerate(sorted([r for r in rows if r["done"] == 0], key=lambda x: -x["pop"]), start=1):
        ws3.append([i, r["pref"], r["name"], r["pop"], r["code"]])
        ws3.cell(row=ws3.max_row, column=4).number_format = "#,##0"

    ws4 = wb.create_sheet("あと一歩（人口順）")
    head(ws4, ["順位", "都道府県", "市区町村", "人口", "空き状況", "お金記事", "足りないもの", "slug"],
         [7, 11, 20, 12, 11, 11, 20, 18])
    near = sorted([r for r in rows if r["sim"] == "○" and r["done"] < 3], key=lambda x: -x["pop"])
    for i, r in enumerate(near, start=1):
        lack = "・".join([n for n, v in (("空き状況", r["vac"]), ("お金記事", r["money"])) if v != "○"])
        ws4.append([i, r["pref"], r["name"], r["pop"], r["vac"], r["money"], lack, r["slug"]])
        ws4.cell(row=ws4.max_row, column=4).number_format = "#,##0"

    ws5 = wb.create_sheet("メモ")
    ws5.column_dimensions["A"].width = 100
    ws5.append(["この表について"])
    ws5["A1"].font = Font(bold=True, size=12)
    for line in [
        "",
        f"人口: 総務省「{source_label}」の「人口（計）」",
        f"      {INDEX_URL}",
        "",
        "行の単位: 市・町・村・特別区。政令指定都市は市単位（行政区の行は除外）。",
        "          郡の合計行と都道府県の合計行は除外。",
        "",
        "シミュレーター: src/lib/data/<slug>.ts（点数の基準を持っている自治体）",
        "空き状況: src/lib/vacancy/index.ts の registry に登録されている自治体",
        "お金記事: src/lib/insurance/city-<slug>.ts（自治体ごとの子育てのお金の記事）",
        "",
        "作り直すとき: python scripts/coverage-report.py <出力先.xlsx>",
        "",
        "突き合わせできなかったサイト側のデータ（あれば名前か県の書き方がずれている）:",
    ]:
        ws5.append([line])
    if unmatched:
        ws5.append(["slug", "名前", "都道府県"])
        for u in unmatched:
            ws5.append(list(u))
    else:
        ws5.append(["なし"])

    wb.save(out_path)


def main():
    if len(sys.argv) != 2:
        fail("使い方: python scripts/coverage-report.py <出力先.xlsx>")
    out_path = sys.argv[1]
    path, label = download_population()
    munis = load_municipalities(path)
    sim, money, vac, vac_only = load_coverage()
    rows, unmatched = build_rows(munis, sim, money, vac, vac_only)
    write_book(out_path, rows, unmatched, label)

    done3 = sum(1 for r in rows if r["done"] == 3)
    print(f"書き出し: {out_path}")
    print(f"市区町村 {len(rows)} / シミュレーター {sum(1 for r in rows if r['sim'] == '○')}"
          f" / 空き状況 {sum(1 for r in rows if r['vac'] == '○')}"
          f" / お金記事 {sum(1 for r in rows if r['money'] == '○')}"
          f" / 3つとも {done3} / 未着手 {sum(1 for r in rows if r['done'] == 0)}")
    if unmatched:
        print(f"※ 突き合わせできなかったデータが {len(unmatched)} 件あります: {unmatched[:5]}")


main()
