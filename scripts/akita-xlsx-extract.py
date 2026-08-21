"""
秋田市の「教育・保育施設等受入可能状況」Excelから表を抜き出す

実行: python scripts/akita-xlsx-extract.py <xlsx>
出力: 標準出力にJSON（fetch-akita-vacancy.ts から呼ぶ）

## 表の作り
- 1シート。「地区／サービス種類／事業所名／0歳〜5歳」の9列
- 空きは人数。0も入っている
- **保育を実施していない歳児は0ではなく斜線**で示されるので、
  セルの罫線（対角線）を見て切り分ける
- 「網掛部分は受入できません」という注記もあるので、塗りつぶしも見て返す
"""

import json
import re
import sys

import openpyxl

AGE_COUNT = 6
HEADER_ROW = 3
COL_WARD = 0
COL_KIND = 1
COL_NAME = 2
COL_AGE0 = 3


def fail(message):
    raise SystemExit(f"[中断] {message}")


def text(value):
    if value is None:
        return ""
    return "".join(str(value).split())


def extract(path):
    book = openpyxl.load_workbook(path)
    sheet = book[book.sheetnames[0]]

    title = text(sheet["A1"].value)
    z = str.maketrans("０１２３４５６７８９", "0123456789")
    m = re.search(r"令和([０-９\d]+)年([０-９\d]+)月([０-９\d]+)日現在", title.translate(z))
    if not m:
        fail(f"基準日を読み取れませんでした: 「{title}」")
    as_of = tuple(int(g) for g in m.groups())

    header = [text(c.value) for c in sheet[HEADER_ROW]][: COL_AGE0 + AGE_COUNT]
    expected = ["地区", "サービス種類", "事業所名"] + [f"{i}歳" for i in range(AGE_COUNT)]
    if [h.translate(z) for h in header] != expected:
        fail(f"見出しが{header}になっています")

    rows = []
    notes = []
    for row in sheet.iter_rows(min_row=HEADER_ROW + 1):
        name = text(row[COL_NAME].value)
        if not name:
            # 表の下にある注意書き
            note = text(row[COL_WARD].value)
            if note:
                notes.append(str(row[COL_WARD].value).strip())
            continue
        cells = row[COL_AGE0 : COL_AGE0 + AGE_COUNT]
        rows.append(
            {
                "ward": text(row[COL_WARD].value),
                "kind": text(row[COL_KIND].value),
                "name": name,
                "values": [c.value for c in cells],
                # 斜線＝保育を実施していない、網掛＝受入できません
                "slashed": [bool(c.border.diagonalDown or c.border.diagonalUp) for c in cells],
                "shaded": [c.fill.patternType is not None for c in cells],
            }
        )

    if not rows:
        fail("受入可能状況の表を取り出せませんでした")

    return {"asOf": as_of, "title": title, "notes": notes, "rows": rows}


def main():
    paths = sys.argv[1:]
    if len(paths) != 1:
        fail("Excelのパスを1つ指定してください。")
    sys.stdout.reconfigure(encoding="utf-8")
    json.dump(extract(paths[0]), sys.stdout, ensure_ascii=False)


if __name__ == "__main__":
    main()
