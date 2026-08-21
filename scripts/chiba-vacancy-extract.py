"""
千葉市の「認定こども園・保育園等受入状況一覧」ExcelとPDFを読んでJSONで返す

実行: python scripts/chiba-vacancy-extract.py <xlsx> <pdf>
出力: 標準出力にJSON（fetch-chiba-vacancy.ts から呼ぶ）

## 表の作り
- Excel は「施設コード／区／分類／園（事業）名／住所／0歳児〜5歳児」。
  空きは記号（◎＝余裕あり、○＝数名、△＝若干名、×＝空きなし）
- **0歳〜2歳がひとつに結合されているセルがある**（家庭的保育事業など）。
  結合の範囲を見て、同じ値を各クラスに配る
- **凡例にない「0」の記載がある**。そのまま返して、呼び出し側で扱いを決める
- 同じ内容のPDFも出ているので、施設名の並びを突き合わせて取りこぼしを見る
"""

import json
import re
import sys

import openpyxl
import pdfplumber

AGE_COUNT = 6
COL_CODE, COL_WARD, COL_KIND, COL_NAME, COL_ADDRESS = 0, 1, 2, 3, 4
COL_ZERO = 5


def fail(message):
    raise SystemExit(f"[中断] {message}")


def text(v):
    if v is None:
        return ""
    return str(v).replace("\n", "").replace("　", " ").strip()


def read_excel(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    width = COL_ZERO + AGE_COUNT
    grid = [list(row[:width]) for row in ws.iter_rows(values_only=True)]

    # 結合セルの中身は左上にしか入らないので、範囲いっぱいに配る
    merged = 0
    for rng in ws.merged_cells.ranges:
        value = ws.cell(rng.min_row, rng.min_col).value
        if rng.min_row < 4:
            continue
        merged += 1
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                if c <= width and r <= len(grid):
                    grid[r - 1][c - 1] = value

    rows = [[text(v) for v in row] for row in grid]

    title = " ".join(t for t in rows[0] if t)
    m = re.search(r"令和(\d+)年(\d+)月(\d+)日時点", title)
    if not m:
        fail(f"基準日を読み取れませんでした: {title}")
    as_of = (int(m.group(1)), int(m.group(2)), int(m.group(3)))

    # 見出しは3行め（0歳〜5歳）
    head = rows[2][COL_ZERO : COL_ZERO + AGE_COUNT]
    if head != [str(i) for i in range(AGE_COUNT)]:
        fail(f"年齢の見出しが変わりました: {head}")

    facilities = [r for r in rows[3:] if r[COL_CODE] and r[COL_NAME]]
    if not facilities:
        fail("施設の行を取り出せませんでした")
    return {"asOf": as_of, "rows": facilities, "merged": merged}


# 表が始まる目印。ここから後ろだけを数えれば、各ページの前置きにある記号が混ざらない
TABLE_HEAD = "区分類園（事業）名住所0歳児1歳児2歳児3歳児4歳児5歳児"


def read_pdf(path):
    """凡例と、記号の数を数えるための表の部分だけのPDF全文を返す"""
    legend = []
    chunks = []
    with pdfplumber.open(path) as pdf:
        for page_index, page in enumerate(pdf.pages):
            flat = "".join((page.extract_text() or "").split())
            if TABLE_HEAD not in flat:
                fail(f"{page_index + 1}ページめに表の見出しが見つかりません")
            chunks.append(flat.split(TABLE_HEAD, 1)[1])
            if page_index == 0:
                # 「◎・・・余裕あり」「○・・・数名」。凡例の間に別の文が挟まることがある
                for mark, label in re.findall(
                    r"([◎○〇△▲×✕])・{2,}(.+?)(?=[◎○〇△▲×✕]|受入可能状況)", flat
                ):
                    legend.append({"mark": mark, "label": label})
    return "".join(chunks), legend


def main():
    paths = sys.argv[1:]
    if len(paths) != 2:
        fail("Excelのパスと PDFのパスを順に指定してください。")
    sys.stdout.reconfigure(encoding="utf-8")

    excel = read_excel(paths[0])
    pdf_text, legend = read_pdf(paths[1])
    if len(legend) < 4:
        fail(f"記号の凡例を読み取れませんでした（{len(legend)}件）")

    # ExcelにあってPDFにない施設がないか（どちらかの取りこぼしを見つける）。
    # PDFはセルの中で改行された名前の文字順が崩れることがあるので、数件は起こりうる
    missing = [
        r[COL_NAME]
        for r in excel["rows"]
        if "".join(r[COL_NAME].split()) not in pdf_text
    ]

    # 表の部分に出てくる記号の数。Excelの生のセルと突き合わせるために返す
    pdf_marks = {m: pdf_text.count(m) for m in "◎○〇△×"}

    json.dump(
        {
            "asOf": excel["asOf"],
            "merged": excel["merged"],
            "legend": legend,
            "rows": excel["rows"],
            "missingInPdf": missing,
            "pdfMarkCounts": pdf_marks,
        },
        sys.stdout,
        ensure_ascii=False,
    )


if __name__ == "__main__":
    main()
